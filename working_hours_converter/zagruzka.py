import openpyxl

wb = openpyxl.load_workbook('1.xlsx')       #входной файл по форме загрузки - надо переименовывать
#sheet1 = wb["1"]       #непонятный лист? удалить
wb.create_sheet("svd")          #во входном файле создается сводный лист за месяц
sheetoutput = wb["svd"]
lineoutput = 2
table_starts_from_x = 1
table_starts_from_y = 1
table_ends_on_x = 1
table_ends_on_y = 1
blck_cntr = 0
for sheetcntr in range(2, 32):
"""    for y1 in range(1,10):           #определяем начало и конец таблицы на каждом листе
        for x1 in range (1,30):
            if sheetread.cell(row=y1, column=x1).value == "Блок" and blck_cntr == 0:
                table_starts_from_x = x1
                table_starts_from_y = y1
                blck_cntr = 1
            if sheetread.cell(row=y1, column=x1).value == "время работы":
                table_ends_on_x = x1
                table_ends_on_y = y1 """
    sheetread = wb[str(sheetcntr)]
    for y in range(4, 24):
        for x in range(3, 10):
            if sheetread.cell(row=y, column=x).value == None:
                break
            else:
                tmpcellout = sheetoutput.cell(row=lineoutput, column=x)
                tmpcellout.value = sheetread.cell(row=y, column=x).value
                tmpcellout = sheetoutput.cell(row=lineoutput, column=2)
                tmpcellout.value = sheetcntr
        if sheetread.cell(row=y, column=x).value == None:
            break
        else:
            lineoutput += 1
    table_starts_from_x = 1
    table_starts_from_y = 1
    table_end_from_x = 1
    table_end_from_y = 1
wb.save("rd.xlsx")