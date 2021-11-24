import openpyxl
wb = openpyxl.load_workbook('sample_data.xlsx', data_only=True)
#во входном файле создается сводный лист за месяц
wb.create_sheet("svd")
sheetoutput = wb["svd"]
lineoutput = 2
table_starts_from_x = 1
table_starts_from_y = 1
table_ends_on_x = 1
table_ends_on_y = 1
blck_cntr = 0   #отсекатель двойного попадания слова "блок"
for sheetcntr in range(1, 32):
    sheetread = wb[str(sheetcntr)]
    print(sheetread)
    for y1 in range(1, 30):           
        #определяем начало и конец таблицы на каждом листе
        for x1 in range(1, 10):
            # определение начала таблицы
            if sheetread.cell(row=y1, column=x1).value == "Блок" and blck_cntr == 0:
                table_starts_from_x = x1
                table_starts_from_y = y1
                blck_cntr = 1
            # определение конца таблицы
            if sheetread.cell(row=y1, column=x1).value == "время работы":
                table_ends_on_x = x1 + 7
                table_ends_on_y = y1
    blck_cntr = 0
    for y in range(table_starts_from_y+1, table_ends_on_y):
        # если данные закончились, то следующий лист
        if sheetread.cell(row=y, column=table_starts_from_x).value == None:
            break
        # копируем данные в сводный лист
        for x in range(table_starts_from_x, table_ends_on_x):
            tmpcellout = sheetoutput.cell(row=lineoutput, column=x)
            tmpcellout.value = sheetread.cell(row=y, column=x).value
            tmpcellout = sheetoutput.cell(row=lineoutput, column=2)
            tmpcellout.value = sheetcntr
        lineoutput += 1
    print("sheet", sheetread, "done" )
    table_starts_from_x = 1
    table_starts_from_y = 1
    table_ends_on_x = 1
    table_ends_on_y = 1
wb.save("1out.xlsx")